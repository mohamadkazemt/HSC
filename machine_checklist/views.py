from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse

from permissions.utils import permission_required
from .models import Checklist, Question, Answer
from BaseInfo.models import MiningMachine, TypeMachine
from accounts.models import UserProfile
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from datetime import datetime
from django_jalali.db import models as jmodels
import jdatetime
import openpyxl
from openpyxl.utils import get_column_letter
from shift_manager.utils import get_current_shift_and_group  # حذف import اضافی

@permission_required("checklist_form")
def checklist_form_view(request):
    machines = MiningMachine.objects.filter(is_active=True)
    return render(request, 'machine_checklist/checklist_form.html', {'machines': machines})

@permission_required("get_questions")
def get_questions(request, machine_id):
    machine = get_object_or_404(MiningMachine, id=machine_id)
    questions = Question.objects.filter(machine_type=machine.machine_type)
    question_list = []
    for question in questions:
        question_list.append({
            'id': question.id,
            'text': question.text,
            'is_required': question.is_required,
            'question_type': question.question_type,
            'options': question.options,
        })
    return JsonResponse({'questions': question_list})


@csrf_exempt
def submit_checklist(request):
    if request.method == 'POST':
        machine_id = request.POST.get('machine')
        try:
            machine = MiningMachine.objects.get(id=machine_id)
            if request.user and request.user.is_authenticated:
                 current_shift, current_group = get_current_shift_and_group(request.user)
                 if hasattr(request.user, 'userprofile'):
                     shift_group = request.user.userprofile.group
                 else:
                     shift_group = current_group
            else:
                current_shift, current_group = get_current_shift_and_group()
                shift_group = current_group
            checklist = Checklist.objects.create(
                                               user=request.user if request.user.is_authenticated else None,
                                               machine=machine,
                                               shift=current_shift,
                                                shift_group=shift_group
                                               )
        except MiningMachine.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'ماشین مورد نظر یافت نشد.'})

        for key, value in request.POST.items():
            if key.startswith('answer_'):
                question_id = key.split('_')[1]
                question = get_object_or_404(Question, id=question_id)
                description = request.POST.get(f'description_{question_id}')
                if question.question_type == 'text':
                    Answer.objects.create(checklist=checklist, question=question, answer_text=value,
                                          description=description)
                elif question.question_type == 'option':
                    Answer.objects.create(checklist=checklist, question=question, selected_option=value,
                                          description=description)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'متد نامعتبر.'})

@permission_required("checklist_list")
def checklist_list_view(request):
    query = request.GET.get('q')
    from_date_str = request.GET.get('start_date', '')
    to_date_str = request.GET.get('end_date', '')
    group_filter = request.GET.get('group_filter', '')
    shift_filter = request.GET.get('shift_filter', '')
    machine_type_filter = request.GET.get('machine_type_filter', '')

    checklists = Checklist.objects.all().order_by('-date')
    if from_date_str:
        try:
            from_date_parts = list(map(int, from_date_str.split('-')))
            from_date_gregorian = jdatetime.date(from_date_parts[0], from_date_parts[1], from_date_parts[2]).togregorian()
            from_date_gregorian = datetime.combine(from_date_gregorian, datetime.min.time())
            checklists = checklists.filter(date__gte=from_date_gregorian)
        except ValueError:
            pass
    if to_date_str:
        try:
            to_date_parts = list(map(int, to_date_str.split('-')))
            to_date_gregorian = jdatetime.date(to_date_parts[0], to_date_parts[1], to_date_parts[2]).togregorian()
            to_date_gregorian = datetime.combine(to_date_gregorian, datetime.max.time())
            checklists = checklists.filter(date__lte=to_date_gregorian)
        except ValueError:
            pass
    if query:
        checklists = checklists.filter(
            Q(user__username__icontains=query) |
            Q(machine__workshop_code__icontains=query) |
            Q(machine__machine_type__name__icontains=query)
        )
    if group_filter:
        checklists = checklists.filter(user__userprofile__group = group_filter)
    if shift_filter:
        checklists = checklists.filter(shift = shift_filter)
    if machine_type_filter:
       checklists = checklists.filter(machine__machine_type_id = machine_type_filter)

    groups = UserProfile.GROUP_CHOICES
    shifts =  Checklist.objects.values_list('shift', flat=True).distinct()
    machine_types = TypeMachine.objects.all()

    page = request.GET.get('page', 1)
    paginator = Paginator(checklists, 10)  # Show 10 checklists per page
    try:
        checklists = paginator.page(page)
    except PageNotAnInteger:
        checklists = paginator.page(1)
    except EmptyPage:
        checklists = paginator.page(paginator.num_pages)
    return render(request, 'machine_checklist/checklist_list.html', {'checklists': checklists, 'from_date': from_date_str, 'to_date': to_date_str, 'query': query, 'groups': groups, 'shifts': shifts, 'machine_types': machine_types })

@permission_required("checklist_detail")
def checklist_detail_view(request, checklist_id):
    checklist = get_object_or_404(Checklist, id=checklist_id)
    answers = Answer.objects.filter(checklist=checklist)
    return render(request, 'machine_checklist/checklist_detail.html', {'checklist': checklist, 'answers': answers})


@permission_required("export_checklists_excel")
def export_checklists_excel(request):
    query = request.GET.get('q')
    from_date_str = request.GET.get('start_date', '')
    to_date_str = request.GET.get('end_date', '')
    group_filter = request.GET.get('group_filter', '')
    shift_filter = request.GET.get('shift_filter', '')
    machine_type_filter = request.GET.get('machine_type_filter', '')
    checklists = Checklist.objects.all().order_by('-date')
    if from_date_str:
        try:
            from_date_parts = list(map(int, from_date_str.split('-')))
            from_date_gregorian = jdatetime.date(from_date_parts[0], from_date_parts[1], from_date_parts[2]).togregorian()
            from_date_gregorian = datetime.combine(from_date_gregorian, datetime.min.time())
            checklists = checklists.filter(date__gte=from_date_gregorian)
        except ValueError:
            pass
    if to_date_str:
        try:
            to_date_parts = list(map(int, to_date_str.split('-')))
            to_date_gregorian = jdatetime.date(to_date_parts[0], to_date_parts[1], to_date_parts[2]).togregorian()
            to_date_gregorian = datetime.combine(to_date_gregorian, datetime.max.time())
            checklists = checklists.filter(date__lte=to_date_gregorian)
        except ValueError:
            pass
    if query:
        checklists = checklists.filter(
            Q(user__username__icontains=query) |
            Q(machine__workshop_code__icontains=query) |
            Q(machine__machine_type__name__icontains=query)
        )
    if group_filter:
        checklists = checklists.filter(user__userprofile__group = group_filter)
    if shift_filter:
        checklists = checklists.filter(shift = shift_filter)
    if machine_type_filter:
       checklists = checklists.filter(machine__machine_type_id = machine_type_filter)

    # Create Excel workbook
    workbook = openpyxl.Workbook()

    # Group checklists by machine type
    grouped_checklists = {}
    for checklist in checklists:
        machine_type = checklist.machine.machine_type
        if machine_type not in grouped_checklists:
            grouped_checklists[machine_type] = []
        grouped_checklists[machine_type].append(checklist)

    for machine_type, checklists_for_type in grouped_checklists.items():
        sheet = workbook.create_sheet(title=machine_type.name)

        # Add headers
        headers = [
            "کاربر", "کد پرسنلی", "ماشین", "نوع ماشین", "تاریخ", "شیفت", "گروه شیفت"
        ]
        # Fetch all unique questions for headers for this machine type
        questions = Question.objects.filter(machine_type=machine_type)
        question_headers = [question.text for question in questions]
        headers.extend(question_headers)
        headers.extend(["توضیحات " + q.text for q in questions])
        sheet.append(headers)

        # Add data
        for checklist in checklists_for_type:
            row = [
                f"{checklist.user.first_name} {checklist.user.last_name}" if (checklist.user.first_name or checklist.user.last_name) else checklist.user.username,
                checklist.user.userprofile.personnel_code if checklist.user.userprofile.personnel_code else "",
                checklist.machine.workshop_code,
                checklist.machine.machine_type.name,
                checklist.date.strftime("%Y/%m/%d %H:%M:%S"),
                checklist.shift,
                checklist.shift_group,
            ]
            answers = Answer.objects.filter(checklist=checklist)
            answer_dict = {answer.question.id: answer for answer in answers}
            for question in questions:
                answer = answer_dict.get(question.id)
                if answer:
                    if question.question_type == 'text':
                        row.append(answer.answer_text if answer.answer_text else '')
                    elif question.question_type == 'option':
                         row.append(answer.selected_option if answer.selected_option else '')
                else:
                   row.append('')
            for question in questions:
               answer = answer_dict.get(question.id)
               if answer:
                   row.append(answer.description if answer.description else '')
               else:
                   row.append('')
            sheet.append(row)
        for column_letter in range(1, len(headers) + 1):
            column_letter = get_column_letter(column_letter)
            sheet.column_dimensions[column_letter].width = 25

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=checklists.xlsx'
    workbook.save(response)
    return response